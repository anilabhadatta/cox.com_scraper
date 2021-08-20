# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import os
import sqlite3
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import string
import random
from openpyxl import load_workbook
from openpyxl import Workbook
main_path = os.getcwd()


def initiate_driver(main_path):
    desired_capabilities = DesiredCapabilities.CHROME
    desired_capabilities["applicationCacheEnabled"] = False
    options = webdriver.ChromeOptions()
    # options.add_argument('headless')
    options.add_argument("--start-maximized")
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    driver = webdriver.Chrome(executable_path=os.path.join(main_path,"chromedriver.exe"), options=options,
                              desired_capabilities=desired_capabilities)
    driver.set_window_size(1920, 1080)
    return driver


def clear_cache(driver):
    driver.get('chrome://settings/clearBrowserData')
    time.sleep(1)
    actions = ActionChains(driver)
    actions.send_keys(Keys.TAB * 7 + Keys.ENTER)
    actions.perform()


def enter_zip_home_page(driver, zipcode):
    time.sleep(1)
    driver.find_element_by_css_selector("a.pf-trigger.pf-location-trigger").click()
    time.sleep(1)
    WebDriverWait(driver, 60).until(
        EC.visibility_of_element_located((By.XPATH, '//input[@id="pf-zipcode"]'))
    )
    zipbox = driver.find_element_by_xpath('//input[@id="pf-zipcode"]')
    zipbox.click()
    zipbox.clear()
    zipbox.send_keys(zipcode)
    driver.find_element_by_css_selector("input.pf-location-panel-submit-form").click()
    time.sleep(5)


def offers(driver, address_list):
    offer_path = os.getcwd()
    excel_list = []
    for item in address_list:
        zipcode = item[0]
        state = item[1]
        city = item[2]
        address = item[3]

        if address not in os.listdir(os.getcwd()):
            os.mkdir(address)
        os.chdir(address)

        driver.get("https://www.cox.com/business/home.html")
        WebDriverWait(driver, 60).until(
            EC.visibility_of_element_located((By.XPATH, "//a[@class=' pf-trigger pf-location-trigger']"))
        )
        enter_zip_home_page(driver, zipcode)

        driver.get("https://www.cox.com/business/offers/shop-all-offers.html")
        WebDriverWait(driver, 60).until(
            EC.visibility_of_element_located((By.XPATH, "//div[@class='flex-container py-3']"))
        )
        containers = driver.find_elements_by_css_selector("div.flex-container.py-3")
        address_path = os.getcwd()

        offer_count = 0
        offer_list = []
        offer_type = ""
        for container in containers:
            offer_name = container.find_element_by_css_selector("span.fx-font-size-18px").text
            mini_container = container.find_elements_by_css_selector('div.inner-container.px-2')
            print(len(mini_container))
            internet = mini_container[1].find_elements_by_css_selector('span.foggyGrey')
            telephone = mini_container[2].find_elements_by_css_selector('span.foggyGrey')
            tv = mini_container[3].find_elements_by_css_selector('span.foggyGrey')
            print(internet, telephone, tv)
            if offer_name.find('+') != -1:
                offer_type = "Bundle"
                if not tv:
                    offer_type += "2"
                else:
                    offer_type += "1"
                if offer_type not in os.listdir():
                    os.mkdir(offer_type)
                os.chdir(offer_type)
                offer_name = offer_name.replace('+', '').replace('\n', '')
            elif internet:
                offer_type = "Telephone"
                if offer_type not in os.listdir():
                    os.mkdir(offer_type)
                os.chdir(offer_type)
            elif telephone:
                offer_type = "Internet"
                if offer_type not in os.listdir():
                    os.mkdir(offer_type)
                os.chdir(offer_type)
            count = 1
            for folder_name in os.listdir():
                folder_name = folder_name.split('-')
                if folder_name[0] == offer_name:
                    count += 1

            offer_folder_name = offer_name + "-" + str(count)
            if offer_folder_name not in os.listdir():
                os.mkdir(offer_folder_name)
            os.chdir(offer_folder_name)

            driver.execute_script("arguments[0].scrollIntoView(true);", container)
            time.sleep(2)
            container.screenshot(offer_folder_name + ".png")

            driver.execute_script(
                '''document.getElementsByClassName('textHyperlinkWithArrow')[''' + str(offer_count) + '''].click()''')
            time.sleep(2)
            offer_details = driver.find_element_by_css_selector('div.modal.fade.show')
            driver.execute_script("arguments[0].scrollIntoView(true);", offer_details)
            time.sleep(2)
            offer_details.find_element_by_css_selector('div.modal-content').screenshot("OfferDetails.png")
            offer_details.find_element_by_css_selector('button.close').click()

            offer_price_and_link = container.find_elements_by_css_selector(
                'div.pr-0.pl-0.outer-container.col-12.alignSelfStretch.alignItemsStart.col-md-3')[1]
            offer_price = offer_price_and_link.find_element_by_css_selector('span').text
            offer_price = offer_price[:len(offer_price) - 6] + "." + offer_price[len(offer_price) - 6:]

            os.chdir(address_path)
            offer_count += 1
            offer_list.append([offer_type, offer_folder_name, zipcode, state, city, address, offer_name, offer_price])

        print(offer_list)
        excel_list = excel_list + address_page(driver, offer_list, 0, False, offer_path)
        os.chdir(offer_path)
        clear_cache(driver)
    write_to_excel(excel_list, "output.xlsx")


def address_page(driver, offer_list, index, errorcheck, offer_path):
    offer_list_updated = []
    address_path = os.getcwd()
    for idx, offer_data in enumerate(offer_list):
        os.chdir(offer_data[0] + "\\" + offer_data[1])
        driver.get('https://www.cox.com/business/offers/shop-all-offers.html')
        WebDriverWait(driver, 60).until(
            EC.visibility_of_element_located((By.XPATH, "//div[@class='flex-container py-3']"))
        )
        driver.find_elements_by_css_selector("a.btn-transactional.btn-lg.btn-block")[max(idx, index)].click()
        time.sleep(10)
        try:
            address_element = WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.ID, "sa"))
            )
            address_element.click()
            address_element.clear()
            address_element.send_keys(offer_data[5])
            driver.find_element_by_id('unit').click()
        except:
            print("Cannot Click on UNIT box")
            if not errorcheck:
                write_db(offer_data + [idx], offer_path)
            else:
                os.chdir(address_path)
                return offer_list_updated
            offer_list_updated.append(offer_data)
            os.chdir(address_path)
            continue

        try:
            city_element = WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.ID, "city-name"))
            )
            city_element.click()
            city_element.clear()
            city_element.send_keys(offer_data[4])
            zipcode = driver.find_element_by_id('z')
            zipcode.click()
            zipcode.clear()
            zipcode.send_keys(offer_data[2])
            address_details = driver.find_element_by_css_selector('div.container-fluid')
            driver.execute_script("arguments[0].scrollIntoView(true);", address_details)
            time.sleep(2)
            address_details.screenshot('Address.png')
            driver.find_element_by_css_selector('input.btn-primary.see-online-deals.mt-2').click()
            time.sleep(10)
        except:
            print("Cannot Click on Check my Address")
            if not errorcheck:
                write_db(offer_data + [idx], offer_path)
            else:
                os.chdir(address_path)
                return offer_list_updated
            offer_list_updated.append(offer_data)
            os.chdir(address_path)
            continue

        curr_url = driver.current_url
        print(curr_url)
        if curr_url == "https://www.cox.com/business/contact-us/contact-sales-address-failure.html":
            if not errorcheck:
                write_db(offer_data[:8] + [idx], offer_path)
            else:
                os.chdir(address_path)
                return offer_list_updated
            offer_list_updated.append(offer_data)
            os.chdir(address_path)
            continue
        print("Flow 1")
        try:
            WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div.modal.fade.multiple-address-modal.show"))
            )
            driver.execute_script('''document.getElementsByClassName('btn btn-primary continue')[0].click();''')
            print("Clicked Flow 1")
        except:
            print("Skipped Flow 1")
            pass
        print("Flow 2")
        try:
            time.sleep(20)
            driver.execute_script('''document.getElementsByClassName('btn-primary ')[1].click();''')
            print("CLICKED FLOW 2")
        except:
            print("Skipped Flow 2")
            pass

        print("GOING TO BILLING")
        try:
            offer_data = billing(driver, offer_data)
        except:
            print("ERROR IN BILLING")
            if not errorcheck:
                write_db(offer_data[:8] + [idx], offer_path)
            else:
                os.chdir(address_path)
                return offer_list_updated
            pass
        print("CAME BACK FROM BILLING")
        offer_list_updated.append(offer_data)
        os.chdir(address_path)
    return offer_list_updated


def click_input_box_cinfo(inputbox, info):
    inputbox.click()
    inputbox.clear()
    inputbox.send_keys(info)


def download_pdf(driver):
    driver.find_elements_by_xpath("//a[@class='nds-blue-text-regular']")[-1].click()
    WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, "//iframe[@class='popup-link']"))
    )
    driver.switch_to.frame(0)
    WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "button.bare.downloadbutton.uiButton--neutral.uiButton"))
    )
    time.sleep(30)
    driver.find_elements_by_css_selector('button.bare.downloadbutton.uiButton--neutral.uiButton')[1].click()
    driver.command_executor._commands["send_command"] = ("POST", '/session/$sessionId/chromium/send_command')
    params = {'cmd': 'Page.setDownloadBehavior', 'params': {'behavior': 'allow', 'downloadPath': os.getcwd()}}
    driver.execute("send_command", params)
    driver.switch_to.default_content()


def screenshot_order_details(driver, count):
    container = driver.find_elements_by_xpath("//div[@class='orderDetails-container']")[-1]
    driver.execute_script("arguments[0].scrollIntoView(true);", container)
    time.sleep(2)
    container.screenshot("OrderDetails" + count + ".png")


def equipment_and_services(driver, count):
    container = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, "//div[@class='nds-cox-item-grid-w']"))
    )

    WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, "//span[@class='checkmark nds-m-top_xx-small']"))
    )
    driver.find_elements_by_xpath("//span[@class='checkmark nds-m-top_xx-small']")[-1].click()
    driver.find_elements_by_xpath("//span[@class='nds-checkbox_faux']")[0].click()
    driver.find_elements_by_xpath("//span[@class='nds-checkbox_faux']")[1].click()
    driver.execute_script("arguments[0].scrollIntoView(true);", container)
    time.sleep(2)
    container.screenshot('Equipment and Services.png')
    screenshot_order_details(driver, str(count))
    WebDriverWait(driver, 60).until(
        EC.invisibility_of_element_located((By.XPATH, "//div[@class='nds-blue-text-medium']"))
    )
    driver.find_element_by_xpath("//button[@class='nds-button_brand']").click()


def customer_info(driver, count):
    container = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, "//div[@class='nds-cox-item-grid-w create-profile']"))
    )
    first_name = driver.find_element_by_xpath('//input[@class="nds-input first-name-input"]')
    click_input_box_cinfo(first_name, "test")
    last_name = driver.find_element_by_xpath('//input[@class="nds-input last-name-input"]')
    click_input_box_cinfo(last_name, "test")
    email_address = driver.find_element_by_xpath('//input[@class="nds-input email-input"]')
    click_input_box_cinfo(email_address, "test@test.com")
    business_name = driver.find_element_by_xpath('//input[@class="nds-input"]')
    click_input_box_cinfo(business_name, "test")
    phone_number = driver.find_element_by_xpath('//input[@class="nds-input phone-number-input"]')
    click_input_box_cinfo(phone_number, "201-111-1111")
    driver.execute_script("arguments[0].scrollIntoView(true);", container)
    time.sleep(2)
    container.screenshot('Customer Information.png')
    screenshot_order_details(driver, str(count))
    WebDriverWait(driver, 60).until(
        EC.invisibility_of_element_located((By.XPATH, "//div[@class='nds-blue-text-medium']"))
    )
    driver.find_element_by_xpath("//button[@class='nds-button_brand']").click()


def installation_options(driver, count):
    container = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, "//div[@class='nds-cox-item-grid-w']"))
        )
    driver.find_elements_by_xpath("//span[@class='checkmark nds-m-top_xx-small']")[-1].click()
    driver.execute_script("arguments[0].scrollIntoView(true);", container)
    time.sleep(2)
    price_installation_page = driver.find_elements_by_xpath("//span[@class='nds-text-medium nds-float_right']")[0].text
    container.screenshot('Installation Options.png')
    screenshot_order_details(driver, str(count))
    WebDriverWait(driver, 60).until(
        EC.invisibility_of_element_located((By.XPATH, "//div[@class='nds-blue-text-medium']"))
    )
    driver.find_element_by_xpath("//button[@class='nds-button_brand']").click()
    return price_installation_page


def submit_order(driver, count):
    WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, "//span[@class='nds-checkbox_faux']"))
    )
    driver.find_elements_by_xpath("//span[@class='nds-checkbox_faux']")[0].click()
    driver.find_elements_by_xpath("//span[@class='nds-checkbox_faux']")[1].click()
    container = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, "//div[@class='nds-cox-item-grid-w']"))
    )
    driver.execute_script("arguments[0].scrollIntoView(true);", container)
    time.sleep(2)
    container.screenshot('Submit your Order.png')
    price_submit_order = driver.find_elements_by_xpath("//span[@class='nds-text-medium nds-float_right']")[2].text
    screenshot_order_details(driver, str(count))
    return price_submit_order


def no_of_lines(driver):
    container = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, "//div[@class='nds-cox-item-grid-w']"))
    )
    driver.execute_script("arguments[0].scrollIntoView(true);", container)
    time.sleep(2)
    container.screenshot('No of Lines.png')
    WebDriverWait(driver, 60).until(
        EC.invisibility_of_element_located((By.XPATH, "//div[@class='nds-blue-text-medium']"))
    )
    driver.find_element_by_xpath("//button[@class='nds-button_brand']").click()


def feature_and_options(driver, offer_type, count):
    container = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, "//div[@class='nds-cox-item-grid-w']"))
    )
    radiobuttons = driver.find_elements_by_xpath("//span[@class='checkmark nds-m-top_xx-small']")
    if len(radiobuttons) == 4:
        radiobuttons[1].click()
        radiobuttons[2].click()
    elif len(radiobuttons) == 2:
        radiobuttons[0].click()
    # if offer_type == "Telephone":
    #     driver.find_elements_by_xpath("//span[@class='checkmark nds-m-top_xx-small']")[1].click()
    #     driver.find_elements_by_xpath("//span[@class='checkmark nds-m-top_xx-small']")[2].click()
    # else:
    #     driver.find_elements_by_xpath("//span[@class='checkmark nds-m-top_xx-small']")[0].click()
    try:
        driver.find_element_by_xpath("//span[@class='nds-checkbox_faux']").click()
    except:
        pass
    driver.execute_script("arguments[0].scrollIntoView(true);", container)
    time.sleep(2)
    container.screenshot('Features and Options.png')
    screenshot_order_details(driver, str(count))
    WebDriverWait(driver, 60).until(
        EC.invisibility_of_element_located((By.XPATH, "//div[@class='nds-blue-text-medium']"))
    )
    driver.find_element_by_xpath("//button[@class='nds-button_brand']").click()


def tv(driver, count):
    container = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, "//div[@class='nds-cox-item-grid-w']"))
    )
    try:
        container.find_element_by_xpath("//span[@class='nds-checkbox_faux']").click()
    except:
        pass
    driver.execute_script("arguments[0].scrollIntoView(true);", container)
    time.sleep(2)
    container.screenshot('TV Outlet and Services.png')
    screenshot_order_details(driver, str(count))
    WebDriverWait(driver, 60).until(
        EC.invisibility_of_element_located((By.XPATH, "//div[@class='nds-blue-text-medium']"))
    )
    driver.find_element_by_xpath("//button[@class='nds-button_brand']").click()


def billing(driver, offer_data):
    print("Billing")
    price_installation_page, price_submit_order, count = "", "", 1
    if offer_data[0] == "Internet":
        # Equipment and Services
        equipment_and_services(driver, count)
        print("equipment_and_services")
        count += 1

    if offer_data[0] == 'Telephone':
        # No of Lines
        no_of_lines(driver)
        print("no_of_lines")

        # Features and options
        feature_and_options(driver, offer_data[0], count)
        print("feature_and_options")
        count += 1

    if offer_data[0] == 'Bundle1':
        # Equipment and Services
        equipment_and_services(driver, count)
        print("equipment_and_services")
        count += 1

        # No of Lines
        no_of_lines(driver)
        print("no_of_lines")

        # Features and options
        feature_and_options(driver, offer_data[0], count)
        print("feature_and_options")
        count += 1

    elif offer_data[0] == 'Bundle2':
        # Equipment and Services
        equipment_and_services(driver, count)
        print("equipment_and_services")
        count += 1

        # TV
        tv(driver, count)
        count += 1

    # Customer Info
    customer_info(driver, count)
    print("customer_info")
    count += 1

    # Installation Options
    price_installation_page = installation_options(driver, count)
    print("installation_options")
    count += 1

    # Submit your Order
    price_submit_order = submit_order(driver, count)
    print("submit_order")

    offer_data += [price_installation_page, price_submit_order]
    download_pdf(driver)
    print("download_pdf")
    return offer_data


def write_to_excel(excel_list, excel_name):
    if excel_name == "corrected_output.xlsx" and "corrected_output.xlsx" in os.listdir():
        wb = load_workbook(excel_name, read_only=False)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['Zipcode', 'State', 'City', 'Address', 'Offer Name', 'Price in all Offer', 'Price in Installation page', 'Price in Submit order'])
    for item in excel_list:
        ws.append(item[2:])
    wb.save(excel_name)


def read_excel(driver):
    excel_path = path1.get()
    # excel_path = "C:\\Users\\anila\\PycharmProjects\\pythonProject\\input.xlsx"
    wb = load_workbook(excel_path, read_only=False)
    ws = wb.active
    address_list = []
    for row in range(2, ws.max_row + 1):
        zipcode = ws.cell(row, 1).value
        state = ws.cell(row, 2).value
        city = ws.cell(row, 3).value
        address = ws.cell(row, 4).value
        address_list.append([zipcode, state, city, address])
    wb.close()
    offers(driver, address_list)


def write_db(error, offer_path):
    print(error)
    conn = sqlite3.connect(offer_path + "\\" + 'offerlogs.db')
    cursor = conn.cursor()
    cursor.execute("insert into ERRORS values (?, ?, ?, ? ,? ,? ,? ,? ,?)", (
        error[0], error[1], error[2], error[3], error[4], error[5], error[6], error[7], error[8]))
    conn.commit()
    conn.close()


def check_db(driver):
    conn = sqlite3.connect('offerlogs.db')
    cursor = conn.cursor()
    offer_path = os.getcwd()
    offer_list_updated = []
    db_deletion = []
    for idx, rows in enumerate(cursor.execute('SELECT * FROM ERRORS')):
        os.chdir(rows[5])
        driver.get("https://www.cox.com/business/home.html")
        WebDriverWait(driver, 60).until(
            EC.visibility_of_element_located((By.XPATH, "//a[@class=' pf-trigger pf-location-trigger']"))
        )
        enter_zip_home_page(driver, rows[2])

        offer_list = address_page(driver, [list(rows[:-1])], rows[-1], True, offer_path)
        print(offer_list)
        if offer_list:
            offer_list_updated += offer_list
            db_deletion.append(rows[-1])

        os.chdir(offer_path)
        clear_cache(driver)
    for idx in db_deletion:
        sql_delete_query = """DELETE from ERRORS where Idx = ?"""
        cursor.execute(sql_delete_query, (idx,))
        conn.commit()
    write_to_excel(offer_list_updated, "corrected_output.xlsx")
    conn.close()


def main(flag, main_path):
    # Initiate chromedriver
    driver = initiate_driver(main_path)
    print(main_path, flag)
    if flag:
        os.chdir(main_path)
        name = "Offers_" + ''.join(random.choices(string.ascii_uppercase + string.digits, k=10))
        os.mkdir(name)
        os.chdir(name)
        conn = sqlite3.connect('offerlogs.db')
        cursor = conn.cursor()
        cursor.execute("DROP TABLE IF EXISTS ERRORS")
        sql = '''CREATE TABLE ERRORS(
                   OfferType CHAR(20) NOT NULL,
                   OfferFolderName CHAR(20) NOT NULL,
                   ZipCode CHAR(20) NOT NULL,
                   State CHAR(20) NOT NULL,
                   City CHAR(20) NOT NULL,
                   Address CHAR(20) NOT NULL,
                   OfferName CHAR(20) NOT NULL,
                   OfferPrice CHAR(20) NOT NULL,
                   Idx INTEGER NOT NULL
                )'''
        cursor.execute(sql)
        conn.commit()
        conn.close()
        read_excel(driver)
        print("Returned from Read Excel")
    else:
        check_db(driver)
        print("Returned from CheckDB")
    driver.quit()
    messagebox.showinfo("SUCCESS", "DONE")
    

if __name__ == '__main__':

    window = Tk()
    window.title("COX CRAWLER")
    window.geometry("350x150")
    window.configure(bg="white")
    window.attributes("-alpha", 0.90)
    path1 = StringVar()

    Label(window, text="Choose Excel File", bg="white").place(x=0, y=30)
    Entry(window, width=30, bg="lightgrey", textvariable=path1).place(x=110, y=30)
    btn1 = Label(window, text="...", width=4, bg="white")
    btn1.place(x=300, y=30)
    btn1.bind("<Button>", lambda e: path1.set(filedialog.askopenfilename(initialdir=os.getcwd(), title="Select Excel File")))

    Button(window, text="RUN", font="none 11", width=4, command=lambda: main(True, main_path), bg="lightgreen").place(x=60, y=100)
    Button(window, text="FIX ERRORS", font="none 11", width=12, command=lambda: main(False, main_path), bg="lightgreen").place(x=160, y=100)
    window.mainloop()
